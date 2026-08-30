#!/usr/bin/env python3
"""Generate Excel learning material for SMK Class 11."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

wb = openpyxl.Workbook()

# ── Style Definitions ────────────────────────────────────────
title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
result_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

def style_data_cell(ws, row, col, align=center):
    cell = ws.cell(row=row, column=col)
    cell.alignment = align
    cell.border = border
    return cell

def add_title(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center

def auto_width(ws, min_width=10, max_width=30):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


# ══════════════════════════════════════════════════════════════
# SHEET 1: TABEL NILAI SISWA
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Tabel Nilai Siswa"

add_title(ws1, 1, "MATERI EXCEL KELAS 11 SMK — TABEL NILAI SISWA", 10)
ws1.row_dimensions[1].height = 35

# Sub title
ws1.merge_cells("A2:J2")
ws1.cell(row=2, column=1, value="Rumus: SUM, AVERAGE, MAX, MIN, COUNT, COUNTIF").font = Font(italic=True, color="666666")

# Header
headers = ["No", "Nama Siswa", "Bindo", "Binggris", "Matematika", "IPA", "IPS", "Jumlah", "Rata-rata", "Status"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 10)

# Data siswa
siswa = [
    [1, "Budi Santoso", 85, 78, 72, 90, 82],
    [2, "Siti Nurhaliza", 92, 88, 95, 85, 90],
    [3, "Andi Pratama", 65, 70, 58, 72, 68],
    [4, "Rina Wulandari", 78, 82, 80, 88, 75],
    [5, "Dedi Kurniawan", 90, 85, 88, 92, 87],
    [6, "Maya Putri", 55, 60, 45, 62, 58],
    [7, "Rizky Pratama", 88, 90, 92, 86, 91],
    [8, "Lestari Dewi", 72, 75, 68, 78, 70],
    [9, "Fajar Nugroho", 80, 82, 75, 84, 79],
    [10, "Diana Sari", 95, 92, 98, 94, 96],
]

for r, data in enumerate(siswa, 5):
    for c, val in enumerate(data, 1):
        ws1.cell(row=r, column=c, value=val)
        style_data_cell(ws1, r, c)

# Formula Jumlah (SUM)
for r in range(5, 15):
    cell_jumlah = ws1.cell(row=r, column=8)
    cell_jumlah.value = f"=SUM(C{r}:G{r})"
    cell_jumlah.fill = formula_fill
    style_data_cell(ws1, r, 8)
    cell_jumlah.number_format = '#,##0'

# Formula Rata-rata (AVERAGE)
for r in range(5, 15):
    cell_avg = ws1.cell(row=r, column=9)
    cell_avg.value = f"=AVERAGE(C{r}:G{r})"
    cell_avg.fill = formula_fill
    style_data_cell(ws1, r, 9)
    cell_avg.number_format = '#,##0.0'

# Formula Status (IF)
for r in range(5, 15):
    cell_status = ws1.cell(row=r, column=10)
    cell_status.value = f'=IF(I{r}>=75,"Lulus","Tidak Lulus")'
    cell_status.fill = formula_fill
    style_data_cell(ws1, r, 10)

# Summary row
ws1.cell(row=16, column=2, value="TOTAL").font = Font(bold=True)
ws1.cell(row=16, column=3, value="=SUM(C5:C14)").fill = result_fill
ws1.cell(row=16, column=4, value="=SUM(D5:D14)").fill = result_fill
ws1.cell(row=16, column=5, value="=SUM(E5:E14)").fill = result_fill
ws1.cell(row=16, column=6, value="=SUM(F5:F14)").fill = result_fill
ws1.cell(row=16, column=7, value="=SUM(G5:G14)").fill = result_fill
for c in range(3, 8):
    style_data_cell(ws1, 16, c)
    ws1.cell(row=16, column=c).font = Font(bold=True)

ws1.cell(row=17, column=2, value="RATA-RATA").font = Font(bold=True)
ws1.cell(row=17, column=3, value="=AVERAGE(C5:C14)").fill = result_fill
ws1.cell(row=17, column=4, value="=AVERAGE(D5:D14)").fill = result_fill
ws1.cell(row=17, column=5, value="=AVERAGE(E5:E14)").fill = result_fill
ws1.cell(row=17, column=6, value="=AVERAGE(F5:F14)").fill = result_fill
ws1.cell(row=17, column=7, value="=AVERAGE(G5:G14)").fill = result_fill
for c in range(3, 8):
    style_data_cell(ws1, 17, c)
    ws1.cell(row=17, column=c).number_format = '#,##0.0'

ws1.cell(row=18, column=2, value="TERTINGGI").font = Font(bold=True)
ws1.cell(row=18, column=3, value="=MAX(C5:C14)").fill = result_fill
ws1.cell(row=18, column=4, value="=MAX(D5:D14)").fill = result_fill
ws1.cell(row=18, column=5, value="=MAX(E5:E14)").fill = result_fill
ws1.cell(row=18, column=6, value="=MAX(F5:F14)").fill = result_fill
ws1.cell(row=18, column=7, value="=MAX(G5:G14)").fill = result_fill
for c in range(3, 8):
    style_data_cell(ws1, 18, c)

ws1.cell(row=19, column=2, value="TERENDAH").font = Font(bold=True)
ws1.cell(row=19, column=3, value="=MIN(C5:C14)").fill = result_fill
ws1.cell(row=19, column=4, value="=MIN(D5:D14)").fill = result_fill
ws1.cell(row=19, column=5, value="=MIN(E5:E14)").fill = result_fill
ws1.cell(row=19, column=6, value="=MIN(F5:F14)").fill = result_fill
ws1.cell(row=19, column=7, value="=MIN(G5:G14)").fill = result_fill
for c in range(3, 8):
    style_data_cell(ws1, 19, c)

# Legend
ws1.cell(row=21, column=1, value="KETERANGAN RUMUS:").font = Font(bold=True, size=12)
legend = [
    ["A2:C2", "=SUM(C5:G5)", "Menjumlahkan semua nilai"],
    ["A3:C3", "=AVERAGE(C5:G5)", "Menghitung rata-rata"],
    ["A4:C4", "=MAX(C5:G5)", "Mencari nilai tertinggi"],
    ["A5:C5", "=MIN(C5:G5)", "Mencari nilai terendah"],
    ["A6:C6", '=IF(I5>=75,"Lulus","Tidak Lulus")', "Kondisi: Jika rata-rata >= 75 Lulus"],
    ["A7:C7", "=COUNT(C5:C14)", "Menghitung jumlah angka"],
]
for i, (rng, formula, desc) in enumerate(legend, 22):
    ws1.cell(row=i, column=1, value=formula).fill = formula_fill
    ws1.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    ws1.cell(row=i, column=3, value=desc)

auto_width(ws1)


# ══════════════════════════════════════════════════════════════
# SHEET 2: VLOOKUP
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. VLOOKUP")

add_title(ws2, 1, "MATERI: VLOOKUP — Pencarian Data Vertikal", 8)
ws2.row_dimensions[1].height = 35

ws2.merge_cells("A2:H2")
ws2.cell(row=2, column=1, value="Rumus: =VLOOKUP(nilai_cari, tabel_array, kolom_index, [FALSE])").font = Font(italic=True, color="666666")

# Tabel Data Siswa
ws2.cell(row=4, column=1, value="TABEL DATA SISWA").font = Font(bold=True, size=12, color="2F5496")
headers2 = ["NIS", "Nama", "Kelas", "Jurusan", "Jenis Kelamin", "Alamat", "No HP"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=5, column=i, value=h)
style_header_row(ws2, 5, 7)

data_siswa = [
    ["2024001", "Budi Santoso", "XI RPL 1", "Rekayasa Perangkat Lunak", "Laki-laki", "Jl. Merdeka No. 10", "081234567890"],
    ["2024002", "Siti Nurhaliza", "XI RPL 1", "Rekayasa Perangkat Lunak", "Perempuan", "Jl. Sudirman No. 25", "081234567891"],
    ["2024003", "Andi Pratama", "XI TKJ 1", "Teknik Komputer & Jaringan", "Laki-laki", "Jl. Gatot Subroto No. 5", "081234567892"],
    ["2024004", "Rina Wulandari", "XI MM 1", "Multimedia", "Perempuan", "Jl. Pahlawan No. 15", "081234567893"],
    ["2024005", "Dedi Kurniawan", "XI RPL 2", "Rekayasa Perangkat Lunak", "Laki-laki", "Jl. Diponegoro No. 8", "081234567894"],
    ["2024006", "Maya Putri", "XI TKJ 2", "Teknik Komputer & Jaringan", "Perempuan", "Jl. Ahmad Yani No. 20", "081234567895"],
    ["2024007", "Rizky Pratama", "XI MM 2", "Multimedia", "Laki-laki", "Jl. Imam Bonjol No. 12", "081234567896"],
    ["2024008", "Lestari Dewi", "XI RPL 1", "Rekayasa Perangkat Lunak", "Perempuan", "Jl. Sultan Agung No. 3", "081234567897"],
    ["2024009", "Fajar Nugroho", "XI TKJ 1", "Teknik Komputer & Jaringan", "Laki-laki", "Jl. Hayam Wuruk No. 7", "081234567898"],
    ["2024010", "Diana Sari", "XI MM 1", "Multimedia", "Perempuan", "Jl. Veteran No. 18", "081234567899"],
]

for r, data in enumerate(data_siswa, 6):
    for c, val in enumerate(data, 1):
        ws2.cell(row=r, column=c, value=val)
        style_data_cell(ws2, r, c)

# Contoh VLOOKUP
ws2.cell(row=18, column=1, value="CONTOH VLOOKUP:").font = Font(bold=True, size=12, color="2F5496")

examples = [
    ["Cari Nama dari NIS", "2024003", "=VLOOKUP(B19,$A$5:$G$14,2,FALSE)", "Andi Pratama"],
    ["Cari Kelas dari NIS", "2024005", "=VLOOKUP(B20,$A$5:$G$14,3,FALSE)", "XI RPL 2"],
    ["Cari Jurusan dari NIS", "2024007", "=VLOOKUP(B21,$A$5:$G$14,4,FALSE)", "Multimedia"],
    ["Cari Alamat dari NIS", "2024001", "=VLOOKUP(B22,$A$5:$G$14,6,FALSE)", "Jl. Merdeka No. 10"],
    ["Cari No HP dari NIS", "2024010", "=VLOOKUP(B23,$A$5:$G$14,7,FALSE)", "081234567899"],
]

ws2.cell(row=19, column=1, value="Soal").font = Font(bold=True)
ws2.cell(row=19, column=2, value="NIS Dicari").font = Font(bold=True)
ws2.cell(row=19, column=3, value="Rumus VLOOKUP").font = Font(bold=True)
ws2.cell(row=19, column=4, value="Hasil").font = Font(bold=True)
style_header_row(ws2, 19, 4)

for i, (soal, nis, formula, hasil) in enumerate(examples, 20):
    ws2.cell(row=i, column=1, value=soal)
    ws2.cell(row=i, column=2, value=nis)
    cell = ws2.cell(row=i, column=3)
    cell.value = formula
    cell.fill = formula_fill
    ws2.cell(row=i, column=4, value=hasil).fill = result_fill
    for c in range(1, 5):
        style_data_cell(ws2, i, c)

# Penjelasan
ws2.cell(row=26, column=1, value="PENJELASAN:").font = Font(bold=True, size=12, color="2F5496")
ws2.merge_cells("A27:H27")
ws2.cell(row=27, column=1, value="=VLOOKUP(nama_koordinat, $A$5:$G$14, nomor_kolom, FALSE)").font = Font(bold=True)
ws2.merge_cells("A28:H28")
ws2.cell(row=28, column=1, value="$ = tanda absolut (agar range tidak berpindah saat di-drag)").font = Font(italic=True)
ws2.merge_cells("A29:H29")
ws2.cell(row=29, column=1, value="FALSE = exact match (cocok persis)").font = Font(italic=True)
ws2.merge_cells("A30:H30")
ws2.cell(row=30, column=1, value="TRUE = approximate match (terdekat, data harus terurut)").font = Font(italic=True)

auto_width(ws2)


# ══════════════════════════════════════════════════════════════
# SHEET 3: IF & NESTED IF
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. IF & Nested IF")

add_title(ws3, 1, "MATERI: IF & Nested IF — Percabangan Kondisional", 8)
ws3.row_dimensions[1].height = 35

ws3.merge_cells("A2:H2")
ws3.cell(row=2, column=1, value='Rumus: =IF(kondisi, "nilai_jika_benar", "nilai_jika_salah")').font = Font(italic=True, color="666666")

# Tabel Nilai
headers3 = ["No", "Nama", "Nilai UTS", "Nilai UAS", "Nilai Tugas", "Rata-rata", "Grade", "Keterangan"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, 8)

data_nilai = [
    [1, "Budi", 85, 88, 80],
    [2, "Siti", 92, 95, 90],
    [3, "Andi", 55, 60, 50],
    [4, "Rina", 78, 82, 75],
    [5, "Dedi", 90, 92, 88],
    [6, "Maya", 45, 50, 40],
    [7, "Rizky", 88, 85, 92],
    [8, "Lestari", 70, 72, 68],
]

for r, data in enumerate(data_nilai, 5):
    for c, val in enumerate(data, 1):
        ws3.cell(row=r, column=c, value=val)
        style_data_cell(ws3, r, c)
    
    # Rata-rata
    cell = ws3.cell(row=r, column=6)
    cell.value = f"=AVERAGE(C{r}:E{r})"
    cell.fill = formula_fill
    cell.number_format = '#,##0.0'
    style_data_cell(ws3, r, 6)
    
    # Grade (Nested IF)
    cell_g = ws3.cell(row=r, column=7)
    cell_g.value = f'=IF(F{r}>=90,"A",IF(F{r}>=80,"B",IF(F{r}>=70,"C",IF(F{r}>=60,"D","E"))))'
    cell_g.fill = formula_fill
    style_data_cell(ws3, r, 7)
    
    # Keterangan
    cell_k = ws3.cell(row=r, column=8)
    cell_k.value = f'=IF(F{r}>=75,"Lulus","Tidak Lulus")'
    cell_k.fill = formula_fill
    style_data_cell(ws3, r, 8)

# Summary
ws3.cell(row=14, column=1, value="CONTOH RUMUS IF:").font = Font(bold=True, size=12, color="2F5496")
examples_if = [
    ['=IF(A1>=70,"Lulus","Tidak Lulus")', "Kondisi sederhana: jika >= 70 Lulus"],
    ['=IF(A1>=90,"A",IF(A1>=80,"B","C"))', "Nested IF: Grade A, B, atau C"],
    ['=IF(AND(A1>=70,B1>=70),"Lulus","Remedial")', "AND: kedua kondisi harus benar"],
    ['=IF(OR(A1>=90,B1>=90),"Juara","Biasa")', "OR: salah satu kondisi benar"],
    ['=IF(ISBLANK(A1),"Kosong",A1)', "ISBLANK: cek sel kosong"],
]
for i, (rumus, keterangan) in enumerate(examples_if, 15):
    ws3.cell(row=i, column=1, value=rumus).fill = formula_fill
    ws3.merge_cells(start_row=i, start_column=4, end_row=i, end_column=8)
    ws3.cell(row=i, column=4, value=keterangan)

auto_width(ws3)


# ══════════════════════════════════════════════════════════════
# SHEET 4: SUMIF & COUNTIF
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. SUMIF & COUNTIF")

add_title(ws4, 1, "MATERI: SUMIF & COUNTIF — Penjumlahan & Hitung Bersyarat", 8)
ws4.row_dimensions[1].height = 35

# Data Penjualan
ws4.cell(row=3, column=1, value="DATA PENJUALAN TOKO").font = Font(bold=True, size=12, color="2F5496")
headers4 = ["No", "Tanggal", "Barang", "Kategori", "Qty", "Harga Satuan", "Total"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, 7)

data_jual = [
    [1, "01/06/2026", "Pensil 2B", "Alat Tulis", 10, 3000],
    [2, "01/06/2026", "Buku Tulis", "Alat Tulis", 20, 5000],
    [3, "01/06/2026", "Snack Cemilan", "Makanan", 15, 8000],
    [4, "02/06/2026", "Pulpen Pilot", "Alat Tulis", 8, 7000],
    [5, "02/06/2026", "Mineral Water", "Minuman", 30, 4000],
    [6, "02/06/2026", "Kopi Sachet", "Minuman", 25, 2500],
    [7, "03/06/2026", "Kertas HVS", "Alat Tulis", 5, 25000],
    [8, "03/06/2026", "Roti Bakar", "Makanan", 12, 10000],
    [9, "03/06/2026", "Teh Botol", "Minuman", 20, 5000],
    [10, "03/06/2026", "Spidol Warna", "Alat Tulis", 6, 15000],
]

for r, data in enumerate(data_jual, 5):
    for c, val in enumerate(data, 1):
        ws4.cell(row=r, column=c, value=val)
        style_data_cell(ws4, r, c)
    
    # Total Qty x Harga
    cell = ws4.cell(row=r, column=7)
    cell.value = f"=E{r}*F{r}"
    cell.fill = formula_fill
    cell.number_format = '#,##0'
    style_data_cell(ws4, r, 7)

# SUMIF & COUNTIF
ws4.cell(row=17, column=1, value="RUMUS SUMIF & COUNTIF:").font = Font(bold=True, size=12, color="2F5496")

formulas = [
    ["Total Alat Tulis", '=SUMIF(D5:D14,"Alat Tulis",G5:G14)', '=COUNTIF(D5:D14,"Alat Tulis")'],
    ["Total Makanan", '=SUMIF(D5:D14,"Makanan",G5:G14)', '=COUNTIF(D5:D14,"Makanan")'],
    ["Total Minuman", '=SUMIF(D5:D14,"Minuman",G5:G14)', '=COUNTIF(D5:D14,"Minuman")'],
    ["Total Semua", "=SUM(G5:G14)", "=COUNT(G5:G14)"],
]

ws4.cell(row=18, column=1, value="Kategori").font = Font(bold=True)
ws4.cell(row=18, column=2, value="Total Penjualan (SUMIF)").font = Font(bold=True)
ws4.cell(row=18, column=3, value="Jumlah Transaksi (COUNTIF)").font = Font(bold=True)
style_header_row(ws4, 18, 3)

for i, (kat, sumif, countif) in enumerate(formulas, 19):
    ws4.cell(row=i, column=1, value=kat)
    ws4.cell(row=i, column=2, value=sumif).fill = formula_fill
    ws4.cell(row=i, column=3, value=countif).fill = formula_fill
    for c in range(1, 4):
        style_data_cell(ws4, i, c)
        if c == 2:
            ws4.cell(row=i, column=c).number_format = '#,##0'

auto_width(ws4)


# ══════════════════════════════════════════════════════════════
# SHEET 5: CHART
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Chart Grafik")

add_title(ws5, 1, "MATERI: Membuat Chart / Grafik di Excel", 6)
ws5.row_dimensions[1].height = 35

# Data untuk chart
ws5.cell(row=3, column=1, value="DATA PENJUALAN BULANAN").font = Font(bold=True, size=12, color="2F5496")
headers5 = ["Bulan", "Alat Tulis", "Makanan", "Minuman", "Elektronik"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=4, column=i, value=h)
style_header_row(ws5, 4, 5)

data_chart = [
    ["Januari", 2500000, 3200000, 1800000, 5000000],
    ["Februari", 2800000, 3500000, 2100000, 4800000],
    ["Maret", 3100000, 3800000, 2500000, 5200000],
    ["April", 2900000, 4100000, 2800000, 5500000],
    ["Mei", 3200000, 4500000, 3200000, 5800000],
    ["Juni", 3500000, 4800000, 3500000, 6100000],
]

for r, data in enumerate(data_chart, 5):
    for c, val in enumerate(data, 1):
        ws5.cell(row=r, column=c, value=val)
        style_data_cell(ws5, r, c)
        if c > 1:
            ws5.cell(row=r, column=c).number_format = '#,##0'

# Bar Chart
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Penjualan Bulanan per Kategori"
chart1.y_axis.title = "Rupiah (Rp)"
chart1.x_axis.title = "Bulan"
chart1.width = 25
chart1.height = 15

data_ref = Reference(ws5, min_col=2, max_col=5, min_row=4, max_row=10)
cats = Reference(ws5, min_col=1, min_row=5, max_row=10)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws5.add_chart(chart1, "A13")

# Pie Chart
pie_data = Reference(ws5, min_col=2, max_col=5, min_row=4, max_row=4)
pie_vals = Reference(ws5, min_col=2, max_col=5, min_row=11, max_row=11)
ws5.cell(row=11, column=1, value="TOTAL")
for c in range(2, 6):
    col_letter = get_column_letter(c)
    ws5.cell(row=11, column=c, value=f"=SUM({col_letter}5:{col_letter}10)")
    ws5.cell(row=11, column=c).fill = result_fill
    style_data_cell(ws5, 11, c)
    ws5.cell(row=11, column=c).number_format = '#,##0'

pie = PieChart()
pie.title = "Komposisi Penjualan per Kategori"
pie.width = 18
pie.height = 12
pie.add_data(Reference(ws5, min_col=2, max_col=5, min_row=10, max_row=11), titles_from_data=True)
pie.set_categories(Reference(ws5, min_col=2, max_col=5, min_row=4, max_row=4))
ws5.add_chart(pie, "A30")

auto_width(ws5)


# ══════════════════════════════════════════════════════════════
# SHEET 6: LATIHAN
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Latihan")

add_title(ws6, 1, "LATIHAN — Kerjakan di Excel!", 6)
ws6.row_dimensions[1].height = 35

ws6.merge_cells("A2:F2")
ws6.cell(row=2, column=1, value="Isi kolom yang kosong dengan rumus yang benar!").font = Font(bold=True, color="CC0000", size=12)

# Latihan 1
ws6.cell(row=4, column=1, value="LATIHAN 1: Hitunglah!").font = Font(bold=True, size=12, color="2F5496")
headers6 = ["No", "Nama", "Bindo", "Binggris", "Matematika", "Jumlah", "Rata-rata"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=5, column=i, value=h)
style_header_row(ws6, 5, 7)

latihan1 = [
    [1, "Ahmad", 75, 80, 70],
    [2, "Bunga", 88, 92, 85],
    [3, "Citra", 60, 65, 55],
    [4, "Doni", 95, 90, 98],
    [5, "Eka", 72, 68, 75],
]
for r, data in enumerate(latihan1, 6):
    for c, val in enumerate(data, 1):
        ws6.cell(row=r, column=c, value=val)
        style_data_cell(ws6, r, c)
    # Kolom Jumlah & Rata-rata kosong untuk diisi siswa
    for c in [6, 7]:
        ws6.cell(row=r, column=c).fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        style_data_cell(ws6, r, c)

ws6.cell(row=12, column=1, value='Petunjuk: Jumlah = SUM(), Rata-rata = AVERAGE()').font = Font(italic=True, color="CC0000")

# Latihan 2
ws6.cell(row=14, column=1, value="LATIHAN 2: VLOOKUP!").font = Font(bold=True, size=12, color="2F5496")
ws6.merge_cells("A14:F14")

ws6.cell(row=15, column=1, value="Cari nama dan kelas berdasarkan NIS menggunakan VLOOKUP!").font = Font(italic=True)

headers_l2 = ["NIS", "Nama", "Kelas"]
for i, h in enumerate(headers_l2, 1):
    ws6.cell(row=16, column=i, value=h)
style_header_row(ws6, 16, 3)

nis_data = [
    ["2024003", "", ""],
    ["2024007", "", ""],
    ["2024001", "", ""],
]
for r, data in enumerate(nis_data, 17):
    for c, val in enumerate(data, 1):
        ws6.cell(row=r, column=c, value=val)
        style_data_cell(ws6, r, c)
    for c in [2, 3]:
        ws6.cell(row=r, column=c).fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

ws6.merge_cells("A20:F20")
ws6.cell(row=20, column=1, value='Petunjuk: =VLOOKUP(A17, $A$5:$G$14, nomor_kolom, FALSE)').font = Font(italic=True, color="CC0000")

# Latihan 3
ws6.cell(row=22, column=1, value="LATIHAN 3: Buatlah Grade dengan Nested IF!").font = Font(bold=True, size=12, color="2F5496")
ws6.merge_cells("A22:F22")

headers_l3 = ["No", "Nama", "Nilai", "Grade"]
for i, h in enumerate(headers_l3, 1):
    ws6.cell(row=23, column=i, value=h)
style_header_row(ws6, 23, 4)

latihan3 = [
    [1, "Fajar", 92],
    [2, "Gita", 78],
    [3, "Hendra", 55],
    [4, "Indah", 85],
    [5, "Joko", 68],
]
for r, data in enumerate(latihan3, 24):
    for c, val in enumerate(data, 1):
        ws6.cell(row=r, column=c, value=val)
        style_data_cell(ws6, r, c)
    ws6.cell(row=r, column=4).fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    style_data_cell(ws6, r, 4)

ws6.merge_cells("A29:F29")
ws6.cell(row=29, column=1, value='Petunjuk: A (>=90), B (>=80), C (>=70), D (>=60), E (<60)').font = Font(italic=True, color="CC0000")

auto_width(ws6)


# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
output = "/root/.hermes/profiles/yangyang/output/Excel_Kelas11_SMK.xlsx"
wb.save(output)
print("File saved: " + output)
print("Sheets: " + str(wb.sheetnames))
